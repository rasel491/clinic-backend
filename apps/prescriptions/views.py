# apps/prescriptions/views.py
from django.db import models
from rest_framework import viewsets, status, mixins
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, JSONParser
from django_filters import rest_framework as filters
from django.db.models import Q, Count, Sum, Avg, F, ExpressionWrapper, DecimalField
from django.utils import timezone
from datetime import datetime, timedelta, date
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
import json
import csv


import os
import json
import csv
import io
from django.http import HttpResponse
from django.template.loader import render_to_string

from core.constants import UserRoles, AuditActions
from .models import (
    Prescription, Medication, PrescriptionItem,
    PrescriptionTemplate, TemplateMedication
)
from .serializers import (
    PrescriptionSerializer, MedicationSerializer,
    PrescriptionItemSerializer, PrescriptionTemplateSerializer,
    TemplateMedicationSerializer, PrescriptionDispenseSerializer,
    PrescriptionRefillSerializer, PrescriptionSearchSerializer,
    MedicationStockUpdateSerializer, PrescriptionStatsSerializer
)
from .filters import PrescriptionFilter, MedicationFilter, PrescriptionTemplateFilter
from .permissions import (
    PrescriptionPermissions, MedicationPermissions,
    PrescriptionTemplatePermissions, CanDispensePrescription,
    CanUpdateStock
)

import logging
logger = logging.getLogger(__name__)

try:
    import xlwt
    HAS_XLWT = True
except ImportError:
    HAS_XLWT = False
    xlwt = None  # Set to None to avoid NameError

try:
    import openpyxl
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None
    Workbook = None
class PrescriptionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Prescription CRUD operations
    """
    queryset = Prescription.objects.select_related(
        'patient', 'patient__user',
        'doctor', 'doctor__user',
        'visit', 'dispensing_pharmacy', 'dispensed_by'
    ).prefetch_related(
        'items', 'items__medication'
    ).filter(deleted_at__isnull=True)
    
    serializer_class = PrescriptionSerializer
    permission_classes = [IsAuthenticated, PrescriptionPermissions]
    filter_backends = [filters.DjangoFilterBackend]
    filterset_class = PrescriptionFilter
    
    def get_queryset(self):
        """Filter queryset based on user role"""
        queryset = super().get_queryset()
        user = self.request.user
        
        # Apply branch scope middleware filter
        
        # Patients can only see their own prescriptions
        if user.role == UserRoles.PATIENT:
            queryset = queryset.filter(patient__user=user)
        
        # Doctors can see prescriptions they created
        elif user.role == UserRoles.DOCTOR:
            queryset = queryset.filter(doctor__user=user)
        
        # Cashiers can see prescriptions for their branch
        elif user.role == UserRoles.CASHIER:
            # Assuming branch is set in middleware
            branch_id = getattr(self.request, 'branch_id', None)
            if branch_id:
                queryset = queryset.filter(
                    Q(dispensing_pharmacy_id=branch_id) |
                    Q(dispensing_pharmacy__isnull=True)
                )
        
        return queryset
    
    def perform_create(self, serializer):
        """Set created_by and updated_by on creation"""
        serializer.save(
            created_by=self.request.user,
            updated_by=self.request.user
        )
    
    def perform_update(self, serializer):
        """Set updated_by on update"""
        serializer.save(updated_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def sign(self, request, pk=None):
        """Sign a prescription digitally"""
        prescription = self.get_object()
        
        if prescription.status != 'DRAFT':
            return Response(
                {'error': 'Only draft prescriptions can be signed'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not request.user == prescription.doctor.user:
            return Response(
                {'error': 'Only the prescribing doctor can sign the prescription'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        prescription.status = 'ISSUED'
        prescription.is_signed = True
        prescription.signed_at = timezone.now()
        prescription.save()
        
        # Audit log
        self._log_audit_action(
            request.user,
            AuditActions.UPDATE,
            prescription,
            f"Prescription signed: {prescription.prescription_id}"
        )
        
        return Response({
            'status': 'Prescription signed successfully',
            'prescription_id': prescription.prescription_id,
            'signed_at': prescription.signed_at
        })
    
    @action(detail=True, methods=['post'], permission_classes=[CanDispensePrescription])
    def dispense(self, request, pk=None):
        """Dispense prescription items"""
        prescription = self.get_object()
        serializer = PrescriptionDispenseSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        pharmacy_id = serializer.validated_data['pharmacy_id']
        items_data = serializer.validated_data['items']
        
        # Get pharmacy (branch)
        from apps.clinics.models import Branch
        try:
            pharmacy = Branch.objects.get(id=pharmacy_id)
        except Branch.DoesNotExist:
            return Response(
                {'error': 'Invalid pharmacy ID'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if prescription can be dispensed
        if not prescription.can_be_dispensed:
            return Response(
                {'error': 'Prescription cannot be dispensed'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        dispensed_items = []
        errors = []
        
        with transaction.atomic():
            for item_data in items_data:
                medication_id = item_data['medication_id']
                quantity = item_data['quantity']
                
                # Find prescription item
                try:
                    prescription_item = prescription.items.get(
                        medication_id=medication_id,
                        is_dispensed=False
                    )
                except PrescriptionItem.DoesNotExist:
                    errors.append({
                        'medication_id': medication_id,
                        'error': 'Medication not found in prescription or already dispensed'
                    })
                    continue
                
                # Check stock
                medication = prescription_item.medication
                if medication.current_stock < quantity:
                    errors.append({
                        'medication_id': medication_id,
                        'error': f'Insufficient stock. Available: {medication.current_stock}'
                    })
                    continue
                
                # Dispense item
                prescription_item.dispense(quantity, request.user)
                dispensed_items.append({
                    'medication': medication.name,
                    'quantity': quantity,
                    'unit_price': prescription_item.unit_price,
                    'total': prescription_item.unit_price * quantity
                })
            
            if errors:
                return Response({
                    'dispensed': dispensed_items,
                    'errors': errors
                }, status=status.HTTP_207_MULTI_STATUS)
            
            # Update prescription status
            all_dispensed = prescription.items.filter(is_dispensed=False).count() == 0
            if all_dispensed:
                prescription.status = 'DISPENSED'
                prescription.dispensing_pharmacy = pharmacy
                prescription.dispensed_by = request.user
                prescription.dispensed_at = timezone.now()
            
            prescription.save()
            
            # Audit log
            self._log_audit_action(
                request.user,
                AuditActions.UPDATE,
                prescription,
                f"Prescription dispensed: {prescription.prescription_id}"
            )
        
        return Response({
            'status': 'Prescription dispensed successfully',
            'dispensed_items': dispensed_items,
            'prescription_status': prescription.status
        })
    
    @action(detail=True, methods=['post'])
    def refill(self, request, pk=None):
        """Refill a prescription"""
        prescription = self.get_object()
        serializer = PrescriptionRefillSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        if not prescription.is_refillable:
            return Response(
                {'error': 'This prescription is not refillable'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if prescription.refills_remaining <= 0:
            return Response(
                {'error': 'No refills remaining for this prescription'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            refill_prescription = prescription.refill()
            
            # Audit log
            self._log_audit_action(
                request.user,
                AuditActions.CREATE,
                refill_prescription,
                f"Prescription refilled from: {prescription.prescription_id}"
            )
            
            return Response({
                'status': 'Prescription refilled successfully',
                'refill_prescription_id': refill_prescription.prescription_id,
                'refills_remaining': prescription.refills_remaining
            })
        
        except ValidationError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancel a prescription"""
        prescription = self.get_object()
        
        if prescription.status in ['DISPENSED', 'CANCELLED', 'EXPIRED']:
            return Response(
                {'error': f'Cannot cancel prescription with status: {prescription.status}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        prescription.status = 'CANCELLED'
        prescription.save()
        
        # Audit log
        self._log_audit_action(
            request.user,
            AuditActions.UPDATE,
            prescription,
            f"Prescription cancelled: {prescription.prescription_id}"
        )
        
        return Response({'status': 'Prescription cancelled successfully'})
    
    @action(detail=True, methods=['get'])
    def print(self, request, pk=None):
        """Generate printable prescription"""
        prescription = self.get_object()
        
        # Generate HTML for printing
        context = {
            'prescription': prescription,
            'items': prescription.items.all(),
            'doctor': prescription.doctor,
            'patient': prescription.patient,
            'today': timezone.now().date(),
            'clinic_name': 'Dental Clinic',  # Should come from settings
            'clinic_address': '123 Clinic Street',  # Should come from settings
            'clinic_phone': '+1234567890',  # Should come from settings
        }
        
        html_content = render_to_string('prescriptions/print.html', context)
        
        # Return as PDF (requires reportlab or similar)
        # For now, return HTML that can be printed
        return HttpResponse(html_content, content_type='text/html')
    
    @action(detail=False, methods=['get'])
    def search(self, request):
        """Advanced search for prescriptions"""
        serializer = PrescriptionSearchSerializer(data=request.query_params)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        queryset = self.get_queryset()
        
        # Apply filters
        if patient_id := serializer.validated_data.get('patient_id'):
            queryset = queryset.filter(patient_id=patient_id)
        
        if doctor_id := serializer.validated_data.get('doctor_id'):
            queryset = queryset.filter(doctor_id=doctor_id)
        
        if prescription_id := serializer.validated_data.get('prescription_id'):
            queryset = queryset.filter(prescription_id__icontains=prescription_id)
        
        if status := serializer.validated_data.get('status'):
            queryset = queryset.filter(status=status)
        
        if start_date := serializer.validated_data.get('start_date'):
            queryset = queryset.filter(issue_date__gte=start_date)
        
        if end_date := serializer.validated_data.get('end_date'):
            queryset = queryset.filter(issue_date__lte=end_date)
        
        # Order by latest first
        queryset = queryset.order_by('-issue_date', '-created_at')
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get prescription statistics"""
        queryset = self.get_queryset()
        
        # Calculate statistics
        today = timezone.now().date()
        start_of_month = today.replace(day=1)
        
        total_prescriptions = queryset.count()
        prescriptions_today = queryset.filter(issue_date=today).count()
        prescriptions_this_month = queryset.filter(
            issue_date__gte=start_of_month
        ).count()
        
        # Pending dispensing (issued but not fully dispensed)
        pending_dispensing = queryset.filter(
            status='ISSUED',
            items__is_dispensed=False
        ).distinct().count()
        
        # Revenue calculations
        total_revenue = queryset.aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        
        revenue_today = queryset.filter(
            issue_date=today
        ).aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        
        revenue_this_month = queryset.filter(
            issue_date__gte=start_of_month
        ).aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        
        # Top medications
        top_medications = PrescriptionItem.objects.filter(
            prescription__in=queryset
        ).values(
            'medication__name', 'medication__medicine_id'
        ).annotate(
            total_quantity=Sum('quantity'),
            total_prescriptions=Count('prescription', distinct=True)
        ).order_by('-total_quantity')[:10]
        
        # Prescriptions by status
        prescriptions_by_status = queryset.values('status').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Prescriptions by doctor
        prescriptions_by_doctor = queryset.values(
            'doctor__user__full_name',
            'doctor__doctor_id',
            'doctor__specialization'
        ).annotate(
            count=Count('id'),
            total_revenue=Sum('total_amount')
        ).order_by('-count')[:10]
        
        stats_data = {
            'total_prescriptions': total_prescriptions,
            'prescriptions_today': prescriptions_today,
            'prescriptions_this_month': prescriptions_this_month,
            'pending_dispensing': pending_dispensing,
            'total_revenue': total_revenue,
            'revenue_today': revenue_today,
            'revenue_this_month': revenue_this_month,
            'top_medications': list(top_medications),
            'prescriptions_by_status': dict(prescriptions_by_status.values_list('status', 'count')),
            'prescriptions_by_doctor': list(prescriptions_by_doctor)
        }
        
        serializer = PrescriptionStatsSerializer(stats_data)
        return Response(serializer.data)
    
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export prescriptions data"""
        queryset = self.filter_queryset(self.get_queryset())
        
        format = request.query_params.get('format', 'json')
        
        if format == 'csv':
            return self._export_csv(queryset)
        elif format == 'excel':
            return self._export_excel(queryset)
        else:
            # Default to JSON
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)
            
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        
    def _export_csv(self, queryset):
        """Export prescriptions as CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="prescriptions.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Prescription ID', 'Patient', 'Patient ID', 'Doctor',
            'Issue Date', 'Valid Until', 'Status', 'Type',
            'Diagnosis', 'Total Amount', 'Items Count'
        ])
        
        for prescription in queryset:
            writer.writerow([
                prescription.prescription_id,
                prescription.patient.full_name,
                prescription.patient.patient_id,
                prescription.doctor.full_name,
                prescription.issue_date,
                prescription.valid_until,
                prescription.get_status_display(),
                prescription.get_prescription_type_display(),
                prescription.diagnosis[:100] if prescription.diagnosis else '',
                prescription.total_amount,
                prescription.items.count()
            ])
        
        return response
    
    def _export_excel(self, queryset):
        """Export prescriptions as Excel"""
        if not HAS_XLWT and not HAS_OPENPYXL:
            return Response(
                {'error': 'Excel export requires xlwt or openpyxl library. '
                         'Install with: pip install xlwt openpyxl'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            if HAS_OPENPYXL:
                return self._export_excel_openpyxl(queryset)
            elif HAS_XLWT:
                return self._export_excel_xlwt(queryset)
        except Exception as e:
            logger.error(f"Excel export error: {str(e)}")
            return Response(
                {'error': f'Excel export failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _export_excel_xlwt(self, queryset):
        """Export using xlwt (old format .xls)"""
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="prescriptions.xls"'
        
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Prescriptions')
        
        # Header style
        header_style = xlwt.XFStyle()
        header_style.font.bold = True
        
        # Write headers
        headers = [
            'Prescription ID', 'Patient', 'Patient ID', 'Doctor',
            'Issue Date', 'Valid Until', 'Status', 'Type',
            'Diagnosis', 'Total Amount', 'Items Count'
        ]
        
        for col, header in enumerate(headers):
            ws.write(0, col, header, header_style)
        
        # Write data
        for row, prescription in enumerate(queryset, start=1):
            ws.write(row, 0, prescription.prescription_id)
            ws.write(row, 1, prescription.patient.full_name)
            ws.write(row, 2, prescription.patient.patient_id)
            ws.write(row, 3, prescription.doctor.full_name)
            ws.write(row, 4, prescription.issue_date.strftime('%Y-%m-%d'))
            ws.write(row, 5, prescription.valid_until.strftime('%Y-%m-%d'))
            ws.write(row, 6, prescription.get_status_display())
            ws.write(row, 7, prescription.get_prescription_type_display())
            ws.write(row, 8, prescription.diagnosis[:100] if prescription.diagnosis else '')
            ws.write(row, 9, float(prescription.total_amount))
            ws.write(row, 10, prescription.items.count())
        
        wb.save(response)
        return response
    
    def _export_excel_openpyxl(self, queryset):
        """Export using openpyxl (new format .xlsx)"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="prescriptions.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Prescriptions'
        
        # Write headers
        headers = [
            'Prescription ID', 'Patient', 'Patient ID', 'Doctor',
            'Issue Date', 'Valid Until', 'Status', 'Type',
            'Diagnosis', 'Total Amount', 'Items Count'
        ]
        
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        
        # Write data
        for row, prescription in enumerate(queryset, start=2):
            ws.cell(row=row, column=1, value=prescription.prescription_id)
            ws.cell(row=row, column=2, value=prescription.patient.full_name)
            ws.cell(row=row, column=3, value=prescription.patient.patient_id)
            ws.cell(row=row, column=4, value=prescription.doctor.full_name)
            ws.cell(row=row, column=5, value=prescription.issue_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=6, value=prescription.valid_until.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=7, value=prescription.get_status_display())
            ws.cell(row=row, column=8, value=prescription.get_prescription_type_display())
            ws.cell(row=row, column=9, value=prescription.diagnosis[:100] if prescription.diagnosis else '')
            ws.cell(row=row, column=10, value=float(prescription.total_amount))
            ws.cell(row=row, column=11, value=prescription.items.count())
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(response)
        return response


    
    def _log_audit_action(self, user, action, obj, description):
        """Log audit action"""
        from core.middleware.audit_middleware import AuditContextMiddleware
        
        # Create audit log entry
        AuditContextMiddleware.log_action(
            user=user,
            action=action,
            model_name=obj.__class__.__name__,
            object_id=obj.id,
            description=description,
            ip_address=None,  # Will be set by middleware
            user_agent=None   # Will be set by middleware
        )


class MedicationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Medication CRUD operations
    """
    queryset = Medication.objects.all().order_by('name')
    serializer_class = MedicationSerializer
    permission_classes = [IsAuthenticated, MedicationPermissions]
    filter_backends = [filters.DjangoFilterBackend]
    filterset_class = MedicationFilter
    
    @action(detail=True, methods=['post'], permission_classes=[CanUpdateStock])
    def update_stock(self, request, pk=None):
        """Update medication stock"""
        medication = self.get_object()
        serializer = MedicationStockUpdateSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        action = serializer.validated_data['action']
        quantity = serializer.validated_data['quantity']
        reason = serializer.validated_data.get('reason', '')
        reference_number = serializer.validated_data.get('reference_number', '')
        
        try:
            if action == 'add':
                medication.update_stock(quantity, 'add')
                message = f'Stock increased by {quantity}'
            elif action == 'subtract':
                medication.update_stock(quantity, 'subtract')
                message = f'Stock decreased by {quantity}'
            elif action == 'set':
                medication.current_stock = quantity
                medication.in_stock = quantity > 0
                medication.save(update_fields=['current_stock', 'in_stock', 'updated_at'])
                message = f'Stock set to {quantity}'
            
            # Log stock transaction
            StockTransaction.objects.create(
                medication=medication,
                transaction_type=action.upper(),
                quantity=quantity,
                new_stock=medication.current_stock,
                reason=reason,
                reference_number=reference_number,
                created_by=request.user
            )
            
            return Response({
                'status': 'Stock updated successfully',
                'message': message,
                'current_stock': medication.current_stock,
                'stock_status': medication.stock_status
            })
        
        except ValidationError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        """Get medications with low stock"""
        queryset = self.get_queryset().filter(
            current_stock__gt=0,
            current_stock__lte=F('min_stock_level')
        ).order_by('current_stock')
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def expired(self, request):
        """Get expired medications"""
        from django.utils import timezone
        
        queryset = self.get_queryset().filter(
            expiry_date__lt=timezone.now().date()
        ).order_by('expiry_date')
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def categories(self, request):
        """Get medication categories with counts"""
        categories = Medication.objects.filter(
            is_active=True
        ).values('category').annotate(
            count=Count('id'),
            in_stock=Count('id', filter=Q(in_stock=True))
        ).order_by('category')
        
        return Response(categories)
    
    @action(detail=False, methods=['get'])
    def search_autocomplete(self, request):
        """Search medications for autocomplete"""
        query = request.query_params.get('q', '')
        
        if not query or len(query) < 2:
            return Response([])
        
        medications = Medication.objects.filter(
            Q(name__icontains=query) |
            Q(generic_name__icontains=query) |
            Q(brand__icontains=query) |
            Q(medicine_id__icontains=query)
        ).filter(is_active=True)[:20]
        
        results = []
        for med in medications:
            results.append({
                'id': med.id,
                'medicine_id': med.medicine_id,
                'name': med.name,
                'generic_name': med.generic_name,
                'brand': med.brand,
                'strength': med.strength,
                'form': med.form,
                'unit_price': med.unit_price,
                'current_stock': med.current_stock,
                'requires_prescription': med.requires_prescription,
                'display_text': f"{med.name} ({med.brand}) - {med.strength} - ₹{med.unit_price}"
            })
        
        return Response(results)


class PrescriptionItemViewSet(viewsets.ModelViewSet):
    """
    ViewSet for PrescriptionItem operations
    """
    queryset = PrescriptionItem.objects.select_related(
        'prescription', 'medication', 'dispensed_by'
    ).all()
    
    serializer_class = PrescriptionItemSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Filter queryset based on user role"""
        queryset = super().get_queryset()
        user = self.request.user
        
        # Apply branch scope
        
        # Patients can only see their own prescription items
        if user.role == UserRoles.PATIENT:
            queryset = queryset.filter(prescription__patient__user=user)
        
        # Doctors can see items from their prescriptions
        elif user.role == UserRoles.DOCTOR:
            queryset = queryset.filter(prescription__doctor__user=user)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def dispense_item(self, request, pk=None):
        """Dispense a specific prescription item"""
        item = self.get_object()
        quantity = request.data.get('quantity')
        
        if not quantity or float(quantity) <= 0:
            return Response(
                {'error': 'Valid quantity is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            quantity = float(quantity)
            if quantity > item.remaining_quantity:
                return Response(
                    {'error': f'Cannot dispense more than remaining quantity: {item.remaining_quantity}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check stock
            if item.medication.current_stock < quantity:
                return Response(
                    {'error': f'Insufficient stock. Available: {item.medication.current_stock}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            item.dispense(quantity, request.user)
            
            return Response({
                'status': 'Item dispensed successfully',
                'dispensed_quantity': quantity,
                'remaining_quantity': item.remaining_quantity,
                'is_fully_dispensed': item.is_fully_dispensed
            })
        
        except ValidationError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


class PrescriptionTemplateViewSet(viewsets.ModelViewSet):
    """
    ViewSet for PrescriptionTemplate operations
    """
    queryset = PrescriptionTemplate.objects.select_related(
        'created_by', 'updated_by'
    ).prefetch_related(
        'medications', 'medications__medication'
    ).filter(deleted_at__isnull=True)
    
    serializer_class = PrescriptionTemplateSerializer
    permission_classes = [IsAuthenticated, PrescriptionTemplatePermissions]
    filter_backends = [filters.DjangoFilterBackend]
    filterset_class = PrescriptionTemplateFilter
    
    def perform_create(self, serializer):
        """Set created_by and updated_by on creation"""
        serializer.save(
            created_by=self.request.user,
            updated_by=self.request.user
        )
    
    def perform_update(self, serializer):
        """Set updated_by on update"""
        serializer.save(updated_by=self.request.user)
    
    @action(detail=True, methods=['post'])
    def apply(self, request, pk=None):
        """Apply template to create a new prescription"""
        template = self.get_object()
        
        # Get patient and doctor from request
        patient_id = request.data.get('patient_id')
        doctor_id = request.data.get('doctor_id')
        
        if not patient_id or not doctor_id:
            return Response(
                {'error': 'Patient ID and Doctor ID are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get patient and doctor
        from apps.patients.models import Patient
        from apps.doctors.models import Doctor
        
        try:
            patient = Patient.objects.get(id=patient_id)
        except Patient.DoesNotExist:
            return Response(
                {'error': 'Patient not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            doctor = Doctor.objects.get(id=doctor_id)
        except Doctor.DoesNotExist:
            return Response(
                {'error': 'Doctor not found'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create prescription from template
        with transaction.atomic():
            prescription = Prescription.objects.create(
                prescription_type='NEW',
                patient=patient,
                doctor=doctor,
                diagnosis=template.default_diagnosis,
                notes=template.default_notes,
                instructions=template.default_instructions,
                status='DRAFT',
                created_by=request.user,
                updated_by=request.user
            )
            
            # Add template medications
            for template_med in template.medications.all():
                PrescriptionItem.objects.create(
                    prescription=prescription,
                    medication=template_med.medication,
                    dosage=template_med.default_dosage,
                    frequency=template_med.default_frequency,
                    duration=template_med.default_duration,
                    duration_unit=template_med.default_duration_unit,
                    quantity=template_med.default_quantity,
                    instructions=template_med.default_instructions,
                    unit_price=template_med.medication.unit_price,
                    created_by=request.user,
                    updated_by=request.user
                )
            
            # Increment template usage count
            template.increment_usage()
        
        serializer = PrescriptionSerializer(prescription)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


# Additional models for stock tracking
class StockTransaction(models.Model):
    """Track stock transactions"""
    
    TRANSACTION_TYPES = [
        ('PURCHASE', 'Purchase'),
        ('SALE', 'Sale'),
        ('RETURN', 'Return'),
        ('ADJUSTMENT', 'Adjustment'),
        ('DAMAGE', 'Damage/Loss'),
        ('TRANSFER', 'Transfer'),
    ]
    
    medication = models.ForeignKey(
        Medication,
        on_delete=models.PROTECT,
        related_name='stock_transactions'
    )
    transaction_type = models.CharField(max_length=20, choices=TRANSACTION_TYPES)
    quantity = models.DecimalField(max_digits=10, decimal_places=2)
    previous_stock = models.DecimalField(max_digits=10, decimal_places=2)
    new_stock = models.DecimalField(max_digits=10, decimal_places=2)
    
    reason = models.TextField(blank=True)
    reference_number = models.CharField(max_length=100, blank=True)
    
    created_by = models.ForeignKey(
        'accounts.User',
        on_delete=models.SET_NULL,
        null=True,
        related_name='stock_transactions'
    )
    created_at = models.DateTimeField(auto_now_add=True)
    
    class Meta:
        db_table = 'stock_transactions'
        ordering = ['-created_at']
        indexes = [
            models.Index(fields=['medication', 'created_at']),
            models.Index(fields=['transaction_type']),
        ]
    
    def __str__(self):
        return f"{self.transaction_type} - {self.medication.name} ({self.quantity})"